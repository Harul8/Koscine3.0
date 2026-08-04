"""Two-stage selector + Excel export.

Stage 1 (gate): keep candidates with P(clean) >= a train-calibrated threshold (target
                 clean rate), per side. This caps stop-out FIRST.
Stage 2 (rank):  among survivors, rank by E[ceiling] (quantum) and take the best N/day
                 TOTAL (either side).

Sweeps stop width 0.6-0.9 x ATR and N in {1,2}, for ALL100 / TOP30 / MID_31_100.
Writes a formatted workbook to reports/.
Read-only research (no production runs/ touched).
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from koscine3.data.feature_registry import build_feature_registry  # noqa: E402
from koscine3.data.sources import load_market_data  # noqa: E402
from koscine3.data.universe import UniverseConfig, build_universe  # noqa: E402
from koscine3.datasets.supervised_builder import build_supervised_dataset, model_feature_columns  # noqa: E402
from koscine3.outcomes.clean_move_contract import CleanMoveContract, compute_clean_move_outcomes  # noqa: E402

from lightgbm import LGBMClassifier, LGBMRegressor  # noqa: E402
from sklearn.impute import SimpleImputer  # noqa: E402

TRAIN_END = pd.Timestamp("2023-12-31")
WIDTHS = [0.6, 0.7, 0.8, 0.9]
NS = [1, 2]
GATE_TARGET = 0.75  # train clean rate the Stage-1 gate aims for
OUT = ROOT / "reports" / "clean_move_two_stage_2026-06-11.xlsx"


def _clean(frame, feats):
    return frame[feats].replace([np.inf, -np.inf], np.nan)


def calibrate_gate(p: np.ndarray, y: np.ndarray, target: float) -> float:
    o = np.argsort(-p)
    ps, ys = p[o], y[o]
    cum = np.cumsum(ys) / np.arange(1, len(ys) + 1)
    ok = np.where(cum >= target)[0]
    return float(ps[ok[-1]]) if len(ok) else float(np.quantile(p, 0.99))


def summarize(picks: pd.DataFrame, w: float, days: int, thr: dict) -> dict:
    fd = picks["floor_depth"]
    clean = fd <= w * picks["atr_pct"]
    picks = picks.assign(_clean=clean.astype(int), _year=picks["date"].dt.year)
    so_year = picks.groupby("_year")["_clean"].apply(lambda s: round((1 - s.mean()) * 100, 1)).to_dict()
    return {
        "trades": len(picks),
        "days_traded": int(picks["date"].nunique()),
        "trades_per_day": round(len(picks) / days, 2),
        "stopout_%": round(float((~clean).mean()) * 100, 1),
        "so_2024_%": so_year.get(2024, np.nan),
        "so_2025_%": so_year.get(2025, np.nan),
        "so_2026_%": so_year.get(2026, np.nan),
        "breach_-2%_%": round(float((fd >= 0.02).mean()) * 100, 1),
        "mean_fav_%": round(float(picks["ceiling"].mean()) * 100, 2),
        "mean_fav_clean_%": round(float(picks.loc[clean, "ceiling"].mean()) * 100, 2) if clean.any() else 0.0,
        "med_stop_%": round(float((w * picks["atr_pct"]).median()) * 100, 2),
        "n_long": int((picks["side"] == "long").sum()),
        "n_short": int((picks["side"] == "short").sum()),
        "gate_thr_long": round(thr["long"], 3),
        "gate_thr_short": round(thr["short"], 3),
    }


def main() -> None:
    print("loading + building features ...")
    market = load_market_data()
    registry = build_feature_registry(market)
    universe = build_universe(market, UniverseConfig(cutoff_date="2025-12-31", top_n=100))
    rank = universe.set_index(universe["symbol"].astype(str))["rank"]
    tier_of = rank.apply(lambda r: "TOP30" if r <= 30 else "MID_31_100")

    dataset = build_supervised_dataset(market, universe, registry)
    feats = model_feature_columns(registry, dataset)
    dataset["symbol"] = dataset["symbol"].astype(str)

    outc = compute_clean_move_outcomes(market, universe=universe, contract=CleanMoveContract())
    # entry_open/entry_date already exist on `dataset` (from the swing outcomes); avoid a name clash.
    outc = outc[outc["status"].eq("evaluated")][
        ["date", "symbol", "side", "floor_depth", "ceiling", "atr_pct"]
    ].copy()
    outc["symbol"] = outc["symbol"].astype(str)
    df = dataset.merge(outc, on=["date", "symbol", "side"], how="inner")
    df["tier"] = df["symbol"].map(tier_of)
    train, evl = df[df["date"] <= TRAIN_END], df[df["date"] > TRAIN_END]
    eval_days = int(evl["date"].nunique())
    print(f"train {len(train):,} | eval {len(evl):,} ({eval_days} days) | feats {len(feats)}")

    side_cache = {}
    for side in ("long", "short"):
        tr_s, ev_s = train[train["side"].eq(side)], evl[evl["side"].eq(side)].copy()
        imp = SimpleImputer(strategy="median").fit(_clean(tr_s, feats))
        Xtr, Xev = imp.transform(_clean(tr_s, feats)), imp.transform(_clean(ev_s, feats))
        reg = LGBMRegressor(n_estimators=250, learning_rate=0.05, num_leaves=31, subsample=0.85,
                            colsample_bytree=0.85, random_state=17, verbosity=-1).fit(Xtr, tr_s["ceiling"])
        ev_s["e_ceiling"] = np.maximum(reg.predict(Xev), 0)
        side_cache[side] = (tr_s, ev_s, Xtr, Xev)

    scopes = {"ALL100": None, "TOP30": "TOP30", "MID_31_100": "MID_31_100"}
    summary_rows, trade_rows = [], []
    for w in WIDTHS:
        gated_parts, thr = [], {}
        for side in ("long", "short"):
            tr_s, ev_s, Xtr, Xev = side_cache[side]
            ytr = (tr_s["floor_depth"] <= w * tr_s["atr_pct"]).astype(int)
            clf = LGBMClassifier(n_estimators=250, learning_rate=0.05, num_leaves=31, subsample=0.85,
                                 colsample_bytree=0.85, class_weight="balanced", random_state=17,
                                 verbosity=-1).fit(Xtr, ytr)
            p_tr = clf.predict_proba(Xtr)[:, 1]
            thr[side] = calibrate_gate(p_tr, ytr.to_numpy(), GATE_TARGET)
            e = ev_s.copy()
            e["p_clean"] = clf.predict_proba(Xev)[:, 1]
            gated_parts.append(e[e["p_clean"] >= thr[side]])
        gated = pd.concat(gated_parts, ignore_index=True)
        for scope_name, tier_val in scopes.items():
            pool = gated if tier_val is None else gated[gated["tier"].eq(tier_val)]
            for n in NS:
                picks = pool.sort_values("e_ceiling", ascending=False).groupby("date").head(n)
                summary_rows.append({"scope": scope_name, "width": w, "N_per_day": n,
                                     **summarize(picks, w, eval_days, thr)})
                if scope_name == "ALL100":
                    t = picks.copy()
                    t["width"], t["N_per_day"] = w, n
                    t["clean"] = (t["floor_depth"] <= w * t["atr_pct"]).astype(int)
                    t["stop_level"] = w * t["atr_pct"]
                    trade_rows.append(t)

    summary = pd.DataFrame(summary_rows)
    trades = pd.concat(trade_rows, ignore_index=True)[
        ["width", "N_per_day", "date", "symbol", "side", "tier", "p_clean", "e_ceiling",
         "atr_pct", "stop_level", "entry_date", "entry_open", "floor_depth", "ceiling", "clean"]
    ].sort_values(["width", "N_per_day", "date"])
    print("\nSUMMARY (ALL100):")
    print(summary[summary["scope"].eq("ALL100")].to_string(index=False))

    _write_workbook(summary, trades)
    print(f"\nwrote {OUT}")


def _write_workbook(summary: pd.DataFrame, trades: pd.DataFrame) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    FONT = "Arial"
    head_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    head_fill = PatternFill("solid", fgColor="1F4E78")
    body_font = Font(name=FONT, size=10)
    pct = "0.0%"
    pct2 = "0.00%"

    wb = Workbook()

    # ---- Notes ----
    ws = wb.active
    ws.title = "Notes"
    notes = [
        ("Koscine 3.0 — Clean-Move Two-Stage Selector", True),
        ("", False),
        ("Trade: enter t+1 open, 5-day window. Fixed stop at width x ATR(14)% below entry (above for short); never trailed.", False),
        ("'clean' = stop never hit (exact from daily lows). 'ceiling' = favourable peak vs entry.", False),
        ("", False),
        ("Stage 1 (gate): keep P(clean) >= train-calibrated threshold (target train clean = 75%). Caps stop-out first.", False),
        ("Stage 2 (rank): among survivors, rank by E[ceiling] and take the best N/day TOTAL (either side).", False),
        ("", False),
        ("Universe: top-100 F&O by median turnover_lacs. Tiers: TOP30 (rank 1-30), MID_31_100 (rank 31-100).", False),
        ("Models: LightGBM, trained <= 2023-12-31. Evaluation: 2024-01 .. 2026-05 (out-of-sample).", False),
        ("stopout_% = share of selected trades that hit the stop. breach_-2% = share whose dip reached -2%.", False),
        ("Percentages are realized on out-of-sample evaluation. Reward column mean_fav_clean = avg peak of clean trades.", False),
    ]
    for i, (text, bold) in enumerate(notes, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(name=FONT, bold=bold, size=12 if i == 1 else 10)
    ws.column_dimensions["A"].width = 120

    # ---- Summary ----
    ws = wb.create_sheet("Summary")
    cols = list(summary.columns)
    ws.append(cols)
    for j, _ in enumerate(cols, start=1):
        c = ws.cell(row=1, column=j)
        c.font, c.fill, c.alignment = head_font, head_fill, Alignment(horizontal="center")
    pct_cols = {"stopout_%", "so_2024_%", "so_2025_%", "so_2026_%", "breach_-2%_%",
                "mean_fav_%", "mean_fav_clean_%", "med_stop_%"}
    for _, r in summary.iterrows():
        ws.append([r[c] for c in cols])
    for ri in range(2, ws.max_row + 1):
        for ci, name in enumerate(cols, start=1):
            cell = ws.cell(row=ri, column=ci)
            cell.font = body_font
            if name in pct_cols and isinstance(cell.value, (int, float)):
                cell.value = cell.value / 100.0
                cell.number_format = pct
            if name in {"gate_thr_long", "gate_thr_short"}:
                cell.number_format = "0.000"
    widths = {"scope": 13, "width": 7, "N_per_day": 10}
    for ci, name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(name, 13)
    ws.freeze_panes = "A2"

    # ---- Trades ----
    ws = wb.create_sheet("Trades_ALL100")
    tcols = list(trades.columns)
    ws.append(tcols)
    for j, _ in enumerate(tcols, start=1):
        c = ws.cell(row=1, column=j)
        c.font, c.fill, c.alignment = head_font, head_fill, Alignment(horizontal="center")
    fmt = {"p_clean": "0.000", "e_ceiling": pct2, "atr_pct": pct2, "stop_level": pct2,
           "entry_open": "#,##0.00", "floor_depth": pct2, "ceiling": pct2}
    tr = trades.copy()
    tr["date"] = pd.to_datetime(tr["date"]).dt.strftime("%Y-%m-%d")
    tr["entry_date"] = pd.to_datetime(tr["entry_date"]).dt.strftime("%Y-%m-%d")
    for _, r in tr.iterrows():
        ws.append([r[c] for c in tcols])
    for ri in range(2, ws.max_row + 1):
        for ci, name in enumerate(tcols, start=1):
            cell = ws.cell(row=ri, column=ci)
            cell.font = body_font
            if name in fmt:
                cell.number_format = fmt[name]
    for ci, name in enumerate(tcols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = 12
    ws.freeze_panes = "A2"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)


if __name__ == "__main__":
    main()
