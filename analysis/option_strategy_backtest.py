"""Full options strategy backtest with bucketed selection + diversity + real option P&L.

Universe: top-65 by turnover, two buckets: A=rank1-30, B=rank31-65.
Each day, each bucket: pick the single best (stock, side) by confidence P(big) that passes:
  - tradeable-options gate: 3%-OTM monthly contract with entry premium >= MIN_PREM, OI>0, underlying>=MIN_UND
  - diversity: per-stock 5-day cooldown, <=3/calendar-month, <=6/calendar-quarter
  (walk candidates by confidence; take the first that passes -> slots always fill, stocks rotate)
Buy nearest strike to 3% OTM; real entry premium at t+1; track contract OHLC over the window.
Shows confidence (P(big)) and quantum (pred move). Includes 2026. Writes an Excel.
Read-only research.
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from koscine3.data.feature_registry import build_feature_registry  # noqa: E402
from koscine3.data.sources import load_market_data  # noqa: E402
from koscine3.data.universe import UniverseConfig, build_universe  # noqa: E402
from koscine3.datasets.supervised_builder import build_supervised_dataset, model_feature_columns  # noqa: E402
from koscine3.outcomes.clean_move_contract import CleanMoveContract, compute_clean_move_outcomes  # noqa: E402
from analysis.options_bhavcopy import load_bhavcopy  # noqa: E402

from lightgbm import LGBMClassifier, LGBMRegressor  # noqa: E402
from sklearn.impute import SimpleImputer  # noqa: E402

TRAIN_END = pd.Timestamp("2023-12-31")
BIG_TRAIN = 0.10
OTM = 0.03
MIN_PREM = 5.0      # tradeable-options gate
MIN_UND = 100.0
COOLDOWN = 5
MAX_PER_MONTH = 3
MAX_PER_QTR = 6
OUT = ROOT / "reports" / "option_strategy_top65_buckets_2026-06-11.xlsx"


def _clean(frame, feats):
    return frame[feats].replace([np.inf, -np.inf], np.nan)


def _prem(row) -> float:
    for c in ("open", "close", "settle"):
        v = row[c]
        if pd.notna(v) and v > 0:
            return float(v)
    return np.nan


def main() -> None:
    print("equity + features ...")
    market = load_market_data()
    registry = build_feature_registry(market)
    universe = build_universe(market, UniverseConfig(cutoff_date="2025-12-31", top_n=65))
    rk = universe.set_index(universe["symbol"].astype(str))["rank"]
    bucket_of = {s: ("A_top30" if r <= 30 else "B_next35") for s, r in rk.items()}
    syms = set(rk.index)
    dataset = build_supervised_dataset(market, universe, registry)
    feats = model_feature_columns(registry, dataset)
    dataset["symbol"] = dataset["symbol"].astype(str)

    oc = compute_clean_move_outcomes(market, universe=universe, contract=CleanMoveContract())
    oc = oc[oc["status"].eq("evaluated")][
        ["date", "symbol", "side", "ceiling", "days_to_peak", "entry_date", "entry_open", "window_end_date"]
    ].copy()
    oc["symbol"] = oc["symbol"].astype(str)
    df = dataset.merge(oc, on=["date", "symbol", "side"], how="inner", suffixes=("", "_oc"))
    for c in ("entry_date", "entry_open", "window_end_date"):
        if f"{c}_oc" in df.columns:
            df[c] = df[f"{c}_oc"]
    train, evl = df[df["date"] <= TRAIN_END], df[df["date"] > TRAIN_END].copy()

    print("fit model + score ...")
    for side in ("long", "short"):
        tr_s = train[train["side"].eq(side)]
        imp = SimpleImputer(strategy="median").fit(_clean(tr_s, feats))
        Xtr = imp.transform(_clean(tr_s, feats))
        clf = LGBMClassifier(n_estimators=300, learning_rate=0.05, num_leaves=31, subsample=0.85,
                             colsample_bytree=0.85, class_weight="balanced", random_state=17,
                             verbosity=-1).fit(Xtr, (tr_s["ceiling"] >= BIG_TRAIN).astype(int))
        reg = LGBMRegressor(n_estimators=300, learning_rate=0.05, num_leaves=31, subsample=0.85,
                            colsample_bytree=0.85, random_state=17, verbosity=-1).fit(Xtr, tr_s["ceiling"])
        m = evl["side"].eq(side)
        Xev = imp.transform(_clean(evl[m], feats))
        evl.loc[m, "confidence"] = clf.predict_proba(Xev)[:, 1]
        evl.loc[m, "pred_move"] = np.maximum(reg.predict(Xev), 0)
    evl["opt_type"] = np.where(evl["side"].eq("long"), "CE", "PE")
    evl["bucket"] = evl["symbol"].map(bucket_of)

    cal = np.array(sorted(market["date"].unique()))
    pos = {d: i for i, d in enumerate(cal)}
    win = lambda ed: [pd.Timestamp(cal[j]) for j in range(pos[pd.Timestamp(ed)], min(pos[pd.Timestamp(ed)] + 5, len(cal)))] if pd.Timestamp(ed) in pos else []

    # Option panel for needed dates (top65, near-money).
    close_map = market.set_index(["date", "symbol"])["close"]
    needed = sorted({d for ed in evl["entry_date"].dropna().unique() for d in win(ed)})
    print(f"loading options for {len(needed)} days ...")
    opt = {}
    for k, d in enumerate(needed):
        bc = load_bhavcopy(d)
        if bc.empty:
            continue
        bc = bc[bc["symbol"].isin(syms) & bc["opt_type"].isin(["CE", "PE"])].copy()
        miss = bc["underlying"].isna()
        if miss.any():
            bc.loc[miss, "underlying"] = bc.loc[miss].apply(
                lambda r: close_map.get((pd.Timestamp(d), r["symbol"]), np.nan), axis=1)
        bc = bc[(bc["strike"] >= bc["underlying"] * 0.85) & (bc["strike"] <= bc["underlying"] * 1.15)]
        opt[pd.Timestamp(d)] = bc
        if (k + 1) % 150 == 0:
            print(f"  {k+1}/{len(needed)}")

    def contract_at_entry(p):
        b = opt.get(pd.Timestamp(p["entry_date"]))
        if b is None:
            return None
        ch = b[b["symbol"].eq(p["symbol"]) & b["opt_type"].eq(p["opt_type"])]
        if ch.empty:
            return None
        wend = pd.Timestamp(p["window_end_date"])
        exps = sorted(e for e in ch["expiry"].dropna().unique() if pd.Timestamp(e) >= wend)
        expiry = pd.Timestamp(exps[0]) if exps else pd.Timestamp(sorted(ch["expiry"].dropna().unique())[-1])
        ce = ch[ch["expiry"].eq(expiry)]
        if ce.empty:
            return None
        target = p["entry_open"] * (1 + OTM if p["opt_type"] == "CE" else 1 - OTM)
        krow = ce.iloc[(ce["strike"] - target).abs().argsort().iloc[0]]
        prem = _prem(krow)
        und = krow["underlying"]
        if not (prem and prem >= MIN_PREM) or not (pd.notna(und) and und >= MIN_UND) or not (krow["oi"] > 0):
            return None
        return {"expiry": expiry, "strike": krow["strike"], "entry_prem": prem,
                "entry_vol": int(krow["vol"]) if pd.notna(krow["vol"]) else 0}

    def payoff(p, con):
        highs, t5c, pkc = [], np.nan, np.nan
        dtp = int(p["days_to_peak"]) if pd.notna(p["days_to_peak"]) else 5
        for j, d in enumerate(win(p["entry_date"]), start=1):
            b = opt.get(pd.Timestamp(d))
            if b is None:
                continue
            r = b[b["symbol"].eq(p["symbol"]) & b["opt_type"].eq(p["opt_type"])
                  & b["expiry"].eq(con["expiry"]) & b["strike"].eq(con["strike"])]
            if r.empty:
                continue
            r = r.iloc[0]
            hi = r["high"] if pd.notna(r["high"]) and r["high"] > 0 else r["close"]
            if pd.notna(hi):
                highs.append(float(hi))
            if j == 5 and pd.notna(r["close"]):
                t5c = float(r["close"])
            if j == dtp and pd.notna(r["close"]):
                pkc = float(r["close"])
        if not highs:
            return None
        return max(highs), pkc, t5c

    # Selection with diversity, per bucket per day.
    last_idx, mcount, qcount = {}, {}, {}
    trades = []
    eval_days = sorted(evl["date"].unique())
    for day in eval_days:
        i = pos[pd.Timestamp(day)]
        ym, yq = (day.year, day.month), (day.year, day.quarter)
        for bucket in ("A_top30", "B_next35"):
            cands = evl[(evl["date"].eq(day)) & (evl["bucket"].eq(bucket))].sort_values(
                "confidence", ascending=False)
            for _, p in cands.iterrows():
                s = p["symbol"]
                if i - last_idx.get(s, -10**9) <= COOLDOWN:
                    continue
                if mcount.get((s, ym), 0) >= MAX_PER_MONTH or qcount.get((s, yq), 0) >= MAX_PER_QTR:
                    continue
                if len(win(p["entry_date"])) < 5:
                    continue
                con = contract_at_entry(p)
                if con is None:
                    continue
                pay = payoff(p, con)
                if pay is None:
                    continue
                best, pkc, t5c = pay
                ep = con["entry_prem"]
                trades.append({
                    "date": pd.Timestamp(day), "bucket": bucket, "symbol": s, "side": p["side"],
                    "dir": "UP/Call" if p["side"] == "long" else "DOWN/Put",
                    "confidence": round(float(p["confidence"]), 3), "quantum_pred_%": round(float(p["pred_move"]) * 100, 2),
                    "actual_move_%": round(float(p["ceiling"]) * 100, 2),
                    "strike": con["strike"], "entry_prem": round(ep, 2), "entry_vol": con["entry_vol"],
                    "mult_best": round(best / ep, 2),
                    "mult_peakclose": round(pkc / ep, 2) if pd.notna(pkc) else np.nan,
                    "mult_t5close": round(t5c / ep, 2) if pd.notna(t5c) else np.nan,
                    "year": pd.Timestamp(day).year,
                })
                last_idx[s] = i
                mcount[(s, ym)] = mcount.get((s, ym), 0) + 1
                qcount[(s, yq)] = qcount.get((s, yq), 0) + 1
                break

    res = pd.DataFrame(trades)
    print(f"\nselected trades: {len(res)} | distinct stocks: {res['symbol'].nunique()} "
          f"| trading days: {res['date'].nunique()}")

    def block(d, label):
        out = {"book": label, "n": len(d), "stocks": d["symbol"].nunique()}
        for col in ("mult_best", "mult_peakclose", "mult_t5close"):
            s = d[col].dropna()
            out[f"{col}_mean"] = round(s.mean(), 2)
            out[f"{col}_EV%"] = round((s.clip(0) - 1).mean() * 100, 0)
        out["P>=2x_peak"] = round((d["mult_peakclose"].dropna() >= 2).mean() * 100, 1)
        out["P>=3x_best"] = round((d["mult_best"].dropna() >= 3).mean() * 100, 1)
        return out

    summary = [block(res, "ALL"), block(res[res["bucket"].eq("A_top30")], "A_top30"),
               block(res[res["bucket"].eq("B_next35")], "B_next35")]
    for y in sorted(res["year"].unique()):
        summary.append(block(res[res["year"].eq(y)], f"year {y}"))
    summ = pd.DataFrame(summary)
    print("\n===== SUMMARY (EV = mean option return per trade) =====")
    print(summ.to_string(index=False))
    print("\n===== diversity: top selected stocks =====")
    print(res["symbol"].value_counts().head(12).to_string())
    print("\nsample (highest confidence):")
    print(res.sort_values("confidence", ascending=False).head(10).to_string(index=False))

    _write(summ, res)
    print(f"\nwrote {OUT}")


def _write(summ, res):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    hf, hfill, bf = Font("Arial", bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F4E78"), Font("Arial")
    for name, frame in [("Summary", summ), ("Trades", res.sort_values(["date", "bucket"]))]:
        ws = wb.create_sheet(name) if name != "Summary" else wb.active
        if name == "Summary":
            ws.title = "Summary"
        cols = list(frame.columns)
        ws.append(cols)
        for j in range(1, len(cols) + 1):
            c = ws.cell(1, j); c.font, c.fill, c.alignment = hf, hfill, Alignment(horizontal="center")
        for _, r in frame.iterrows():
            ws.append([pd.Timestamp(r[c]).strftime("%Y-%m-%d") if c == "date" else r[c] for c in cols])
        for ri in range(2, ws.max_row + 1):
            for ci in range(1, len(cols) + 1):
                ws.cell(ri, ci).font = bf
        for ci, c in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = 13 if c not in ("date", "symbol") else 11
        ws.freeze_panes = "A2"
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)


if __name__ == "__main__":
    main()
